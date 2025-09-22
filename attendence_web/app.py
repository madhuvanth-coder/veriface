from flask import Flask, render_template, request, redirect, url_for, session
import openpyxl
import os
from datetime import datetime

app = Flask(__name__)
app.secret_key = "your_secret_key"  # Change to something stronger

# Folder where Excel files are saved
excel_path = "/home/pi/Desktop/file_upload/uploads"

# Dummy teacher login (replace with DB later if needed)
teachers = {
    "teacher1": {"password": "pass123", "class": "ClassA"},
    "teacher2": {"password": "pass456", "class": "ClassB"}
}

# --- Utility: Get latest Excel file ---
def get_latest_excel():
    try:
        files = [f for f in os.listdir(excel_path) if f.endswith(".xlsx")]
        if not files:
            return None
        files = sorted(files, key=lambda f: os.path.getmtime(os.path.join(excel_path, f)))
        return os.path.join(excel_path, files[-1])  # newest file
    except Exception as e:
        print(f"[ERROR] While fetching excel: {e}")
        return None

# --- Utility: Read attendance ---
def read_attendance(teacher_class):
    file = get_latest_excel()
    if not file:
        return [], []  # no file found

    wb = openpyxl.load_workbook(file)
    # Check if the teacher’s sheet/class exists
    if teacher_class not in wb.sheetnames:
        return [], []  # no sheet for this teacher

    sheet = wb[teacher_class]
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))

    headers = data[0] if data else []
    rows = data[1:] if len(data) > 1 else []
    return headers, rows

# --- Routes ---
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        if username in teachers and teachers[username]["password"] == password:
            session["user"] = username
            return redirect(url_for("dashboard"))
        else:
            return render_template("login.html", error="Invalid credentials")

    return render_template("login.html")

@app.route("/dashboard")
def dashboard():
    if "user" not in session:
        return redirect(url_for("login"))

    teacher = session["user"]
    teacher_class = teachers[teacher]["class"]

    headers, data = read_attendance(teacher_class)

    return render_template("dashboard.html", user=teacher, headers=headers, data=data)

@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login"))

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
