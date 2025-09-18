import cv2
from deepface import DeepFace
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# --- Configuration ---
dataset_path = "C:/Users/Acer/Desktop/sih/dataset"
attendance_folder = "C:/Users/Acer/Desktop/sih/attendance"
if not os.path.exists(attendance_folder):
    os.makedirs(attendance_folder)

# Generate today's attendance file name
today = datetime.now().strftime("%Y-%m-%d")
attendance_file = os.path.join(attendance_folder, f"attendance_{today}.xlsx")

# Create Excel file if it doesn't exist
if not os.path.exists(attendance_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Time", "Status"])  # headers
    wb.save(attendance_file)

# Load workbook and sheet
wb = load_workbook(attendance_file)
ws = wb.active

# Track already marked names to avoid duplicates
marked_names = set(ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1))

# --- Haarcascade for face detection ---
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")

# --- Model Config ---
model_name = "ArcFace"

# --- Evaluation Metrics ---
TP, FP, FN, TN = 0, 0, 0, 0  # counters

# Start webcam
cap = cv2.VideoCapture(0)
print("Starting real-time face recognition... (Press 'q' to quit)")

while True:
    ret, frame = cap.read()
    if not ret:
        print("Failed to read from webcam.")
        break

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5)

    if len(faces) == 0:
        cv2.putText(frame, "No face detected", (30, 30),
                    cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
        TN += 1  # No face is a "true negative" in this context

    for (x, y, w, h) in faces:
        face_crop = frame[y:y+h, x:x+w]
        temp_img_path = "temp_face.jpg"
        cv2.imwrite(temp_img_path, face_crop)

        try:
            results = DeepFace.find(
                img_path=temp_img_path,
                db_path=dataset_path,
                model_name=model_name,
                enforce_detection=False
            )

            if len(results[0]) > 0:
                identity_path = results[0].iloc[0]['identity']
                name = os.path.basename(os.path.dirname(identity_path))

                if name not in marked_names:
                    marked_names.add(name)
                    current_time = datetime.now().strftime("%H:%M:%S")
                    ws.append([name, current_time, "Present"])
                    wb.save(attendance_file)
                    print(f"✔ Marked {name} as Present at {current_time}")

                # Metrics update (assuming dataset has only valid faces)
                TP += 1

                # Draw bounding box + name
                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)
                cv2.putText(frame, f"{name}", (x, y - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)
            else:
                FN += 1  # Face detected but not recognized
                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 0, 255), 2)
                cv2.putText(frame, "Unknown", (x, y - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 0, 255), 2)

        except Exception as e:
            print("❌ Error:", e)

        if os.path.exists(temp_img_path):
            os.remove(temp_img_path)

    # Show webcam feed
    cv2.imshow("Face Recognition", frame)

    # Quit
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Cleanup
cap.release()
cv2.destroyAllWindows()
wb.save(attendance_file)

# --- Final Evaluation ---
precision = (TP / (TP + FP) * 100) if (TP + FP) > 0 else 0
recall = (TP / (TP + FN) * 100) if (TP + FN) > 0 else 0
accuracy = ((TP + TN) / (TP + TN + FP + FN) * 100) if (TP + TN + FP + FN) > 0 else 0
prec_frac = precision / 100
recall_frac = recall / 100
f1 = (2 * prec_frac * recall_frac / (prec_frac + recall_frac)) if (prec_frac + recall_frac) > 0 else 0

print("\n📊 Evaluation Metrics:")
print(f"True Positives: {TP}")
print(f"False Positives: {FP}")
print(f"False Negatives: {FN}")
print(f"True Negatives: {TN}")
print(f"Accuracy:  {accuracy:.2f}%")
print(f"Precision: {precision:.2f}%")
print(f"Recall:    {recall:.2f}%")
print(f"F1-Score:  {f1:.2f}%")
print("✅ Evaluation completed. Program exited.")

